#!/usr/bin/env python3

import requests
import urllib3
import sys
import os
from os import path
from lxml import html
from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Font
urllib3.disable_warnings()

# build parser
def parse(business_name_list):
    if response.status_code == 200:
        XPATH_LISTINGS = "//div[@class='search-results organic']//div[@class='v-card']"
        listings = parser.xpath(XPATH_LISTINGS)

        scraped_results = []

        for results in listings:
            XPATH_BUSINESS_NAME = ".//a[@class='business-name']//text()"
            XPATH_TELEPHONE = ".//div[@class='phones phone primary']//text()"
            XPATH_ADDRESS = ".//div[@class='info']//div//p[@itemprop='address']"
            XPATH_STREET = ".//div[@class='street-address']//text()"
            XPATH_CATEGORIES = ".//div[@class='info']//div[contains(@class,'info-section')]//div[@class='categories']//text()"
            XPATH_WEBSITE = ".//div[@class='info']//div[contains(@class,'info-section')]//div[@class='links']//a[contains(@class,'website')]/@href"
            XPATH_LOCALITY = ".//div[@class='info']//div[contains(@class,'info-section')]//div[@class='locality']//text()"
            
            raw_business_name = results.xpath(XPATH_BUSINESS_NAME)
            raw_business_telephone = results.xpath(XPATH_TELEPHONE)
            raw_categories = results.xpath(XPATH_CATEGORIES)
            raw_website = results.xpath(XPATH_WEBSITE)
            raw_street = results.xpath(XPATH_STREET)
            raw_locality = results.xpath(XPATH_LOCALITY)

            business_name = ''.join(raw_business_name).strip() if raw_business_name else None
            telephone = ''.join(raw_business_telephone).strip() if raw_business_telephone else None
            category = ','.join(raw_categories).strip() if raw_categories else None
            website = ''.join(raw_website).strip() if raw_website else None
            street = ''.join(raw_street).strip() if raw_street else None
            locality = ''.join(raw_locality).strip() if raw_locality else None

            if business_name in business_name_list: # this doesn't work as intended
                pass
            else:
                business_details = {
                    'a': category,
                    'b': business_name,
                    'c': telephone,
                    'd': street,
                    'e': locality,
                    'f': website
                }
                business_name_list.append(business_details)
                scraped_results.append(business_details)
            
        return scraped_results

# check if excel file already exists and warn the user
file_exists = path.exists('output.xlsx')

if file_exists == True:
    uhoh = input("\noutput.xlsx already exists! \ntype 'y' to delete and continue or press any other key to abort : ")
    if uhoh == 'y':
        os.remove('output.xlsx')
        print('\ndeleted, continuing...\n')
    else:
        print('\naborting.')
        sys.exit()

# open workbook,sheet
book = Workbook()
sheet = book.active

# apply some styling
#for cell in sheet[1]:
#    cell.font = Font(color="00ffffff")
#    cell.fill = PatternFill(fill_type='solid', bgColor=colors.BLACK)
# set the column headers
sheet.cell(row=1, column=1).value = 'Category'
sheet.cell(row=1, column=2).value = 'Business Name'
sheet.cell(row=1, column=3).value = 'Phone Number'
sheet.cell(row=1, column=4).value = 'Address'
sheet.cell(row=1, column=5).value = 'Locality'
sheet.cell(row=1, column=6).value = 'Website'


# get user input for city / state, abort if user inputs non-conforming state
unencoded_city = input("what's your town? : ")
state = input("and your two letter state name? : ")
if len(state) > 2:
    print('state cannot have more than two letters')
    sys.exit()


# fix spaces in city names
if " " in unencoded_city:
    city = unencoded_city.replace(" ", "-")
else:
    city = unencoded_city

headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           'Accept-Encoding': 'gzip, deflate, br',
           'Accept-Language': 'en-GB,en;q=0.9,en-US;q=0.8,ml;q=0.7',
           'Cache-Control': 'max-age=0',
           'Connection': 'keep-alive',
           'Host': 'www.yellowpages.com',
           'Upgrade-Insecure-Requests': '1',
           'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.140 Safari/537.36'
        }

print('\nbegining... this may take some time to complete.')

# pull keywords page
url = 'http://www.yellowpages.com/{0}-{1}'.format(city, state)
print('retrieving keywords')
response = requests.get(url, verify=False)
if response.status_code == 200:
    print('got response')
    parser = html.fromstring(response.text)

    XPATH_KEYWORDS = "//div[@class='row expand-area']//a//text()"
    keywords = parser.xpath(XPATH_KEYWORDS)
    encode_keywords = [item.replace(' ', '+') for item in keywords]
    keyword_count = len(keywords)
    print('retrieved %s keywords' % (keyword_count))

    print('\n')

    business_name_list = []

    for i in encode_keywords:
        keyword = i

        try:

            url = 'http://www.yellowpages.com/search?search_terms={0}&geo_location_terms={1}%2C+{2}'.format(keyword, city, state)
            unencode_keyword = i.replace('+',' ')
            print("  retrieving:", unencode_keyword)
            print("  --------------------------------")
            response = requests.get(url, verify=False)
            parser = html.fromstring(response.text)
            
            # grab page count
            XPATH_RESULTS = "//div[@class='pagination']//p/child::text()[1]"
            results = parser.xpath(XPATH_RESULTS)
            print("  found: %s results" % (results[0]))
            count_results = int(''.join(results))
            max_per_pg = int(30)
            pages = -(-count_results // max_per_pg)
            print("  total pages:", pages)

            # iterate them
            if pages == 1:
                url = 'http://www.yellowpages.com/search?search_terms={0}&geo_location_terms={1}%2C+{2}'.format(keyword, city, state)
                response = requests.get(url, verify=False)

                #scrape and insert into excel sheet


                scraped_data = parse(business_name_list)
                rows = scraped_data
                for row in rows:
                    sheet.append(row)
                print('  grabbed page')
            
            else:
                while pages > 1:
                    print('  grabbing page: ', pages)
                    url = 'http://www.yellowpages.com/search?search_terms={0}&geo_location_terms={1}%2C+{2}&page={3}'.format(keyword, city, state, pages)
                    response = requests.get(url, verify=False)

                    #scrape and insert into excel sheet
                    scraped_data = parse(business_name_list)
                    rows = scraped_data
                    for row in rows:
                        sheet.append(row)
                    
                    # count down page
                    pages -= 1

                    if pages == 1:
                        print('  grabbing page: ', pages)
                        url = 'http://www.yellowpages.com/search?search_terms={0}&geo_location_terms={1}%2C+{2}'.format(keyword, city, state)
                        response = requests.get(url, verify=False)

                        #scrape and insert into excel sheet
                        scraped_data = parse(business_name_list)
                        rows = scraped_data
                        for row in rows:
                            sheet.append(row)
            
            print('\n')

        except TypeError:
            print('no results', end='\r', flush=True)
        
        except Exception as e:
            print(e)

else:
    print('error: response code: %s' % (response.status_code))

business_name_list.clear()
print('saving to excel file')
book.save('output.xlsx')

'''
# roadmap
employ duplicate skipping <- working on it
'''